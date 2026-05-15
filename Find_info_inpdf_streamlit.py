import re
from pathlib import Path
from io import BytesIO

import fitz  # PyMuPDF
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# =========================
# НАСТРОЙКА СТРАНИЦ
# =========================

def parse_page_selection(page_selection: str | None, total_pages: int) -> list[int]:
    """
    Принимает строки вида:
    1
    1,2
    (1, 2)
    1-3
    (1-3)
    1,5-8
    (1,5-8)

    Возвращает индексы страниц для PyMuPDF, то есть с 0.
    """

    if not page_selection or not page_selection.strip():
        return list(range(total_pages))

    text = page_selection.strip()

    # Убираем скобки: (1, 2) -> 1, 2
    text = text.replace("(", "").replace(")", "")

    # Убираем пробелы
    text = text.replace(" ", "")

    if not text:
        return list(range(total_pages))

    selected_pages = set()

    parts = text.split(",")

    for part in parts:
        if not part:
            continue

        # Диапазон: 1-3
        if "-" in part:
            left, right = part.split("-", 1)

            if not left.isdigit() or not right.isdigit():
                raise ValueError(f"Некорректный диапазон страниц: {part}")

            start_page = int(left)
            end_page = int(right)

            if start_page > end_page:
                start_page, end_page = end_page, start_page

            for page_number in range(start_page, end_page + 1):
                page_index = page_number - 1

                if 0 <= page_index < total_pages:
                    selected_pages.add(page_index)

            continue

        # Одна страница: 5
        if not part.isdigit():
            raise ValueError(f"Некорректный номер страницы: {part}")

        page_number = int(part)
        page_index = page_number - 1

        if 0 <= page_index < total_pages:
            selected_pages.add(page_index)

    return sorted(selected_pages)


# =========================
# PDF
# =========================

def extract_text_from_pdf_file_path(pdf_path: Path, page_selection: str | None) -> tuple[str, str]:
    """
    Извлекает текст из PDF по пути на диске.
    """

    text_parts = []

    try:
        with fitz.open(pdf_path) as doc:
            total_pages = len(doc)
            page_indexes = parse_page_selection(page_selection, total_pages)

            if not page_indexes:
                return "", f"Не найдено подходящих страниц. Всего страниц: {total_pages}"

            for page_index in page_indexes:
                page = doc[page_index]
                text_parts.append(page.get_text("text"))

            pages_note = make_pages_note(page_indexes, total_pages)

    except Exception as e:
        return f"__PDF_READ_ERROR__ {e}", ""

    return "\n".join(text_parts), pages_note


def extract_text_from_uploaded_pdf(uploaded_file, page_selection: str | None) -> tuple[str, str]:
    """
    Извлекает текст из PDF, загруженного через Streamlit uploader.
    """

    text_parts = []

    try:
        file_bytes = uploaded_file.read()

        with fitz.open(stream=file_bytes, filetype="pdf") as doc:
            total_pages = len(doc)
            page_indexes = parse_page_selection(page_selection, total_pages)

            if not page_indexes:
                return "", f"Не найдено подходящих страниц. Всего страниц: {total_pages}"

            for page_index in page_indexes:
                page = doc[page_index]
                text_parts.append(page.get_text("text"))

            pages_note = make_pages_note(page_indexes, total_pages)

    except Exception as e:
        return f"__PDF_READ_ERROR__ {e}", ""

    return "\n".join(text_parts), pages_note


def make_pages_note(page_indexes: list[int], total_pages: int) -> str:
    human_pages = [str(i + 1) for i in page_indexes]
    return f"Обработаны страницы: {', '.join(human_pages)} из {total_pages}"


def normalize_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


# =========================
# ПОИСК ДАННЫХ
# =========================

def clean_value(value: str) -> str:
    value = value.replace("\n", " ")
    value = re.sub(r"\s+", " ", value)
    value = value.strip(" :-–—;,.")
    return value


def find_by_patterns(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            value = clean_value(match.group(1))
            if value:
                return value

    return ""


def extract_developer_name(text: str) -> str:
    patterns = [
        r"Наименование\s+застройщика\s*[:\-–—]?\s*(.+?)(?=\s+(?:ИНН|ОГРН|Адрес|Место\s+нахождения|Наименование\s+объекта|Проектная\s+декларация|Разрешение|$))",
        r"Застройщик\s*[:\-–—]?\s*(.+?)(?=\s+(?:ИНН|ОГРН|Адрес|Место\s+нахождения|Наименование\s+объекта|Проектная\s+декларация|Разрешение|$))",
        r"Полное\s+наименование\s+застройщика\s*[:\-–—]?\s*(.+?)(?=\s+(?:ИНН|ОГРН|Адрес|Место\s+нахождения|Наименование\s+объекта|$))",
    ]

    return find_by_patterns(text, patterns)


def extract_building_area(text: str) -> str:
    patterns = [
        r"Общая\s+площадь\s+здания\s*[:\-–—]?\s*([\d\s.,]+(?:кв\.?\s*м|м2|м²)?)",
        r"Общая\s+площадь\s+объекта\s+капитального\s+строительства\s*[:\-–—]?\s*([\d\s.,]+(?:кв\.?\s*м|м2|м²)?)",
        r"Площадь\s+здания\s*[:\-–—]?\s*([\d\s.,]+(?:кв\.?\s*м|м2|м²)?)",
        r"Общая\s+площадь\s*[:\-–—]?\s*([\d\s.,]+(?:кв\.?\s*м|м2|м²)?)",
    ]

    return find_by_patterns(text, patterns)


def extract_commissioning_date(text: str) -> str:
    date_pattern = r"(\d{1,2}[./-]\d{1,2}[./-]\d{4}|\d{4}-\d{1,2}-\d{1,2})"

    patterns = [
        rf"Дата\s+ввода\s+(?:объекта\s+)?в\s+эксплуатацию\s*[:\-–—]?\s*{date_pattern}",
        rf"Дата\s+ввода\s+объекта\s+капитального\s+строительства\s+в\s+эксплуатацию\s*[:\-–—]?\s*{date_pattern}",
        rf"Дата\s+выдачи\s+разрешения\s+на\s+ввод\s+(?:объекта\s+)?в\s+эксплуатацию\s*[:\-–—]?\s*{date_pattern}",
        rf"Разрешение\s+на\s+ввод\s+(?:объекта\s+)?в\s+эксплуатацию.+?\sот\s+{date_pattern}",
    ]

    return find_by_patterns(text, patterns)


def process_pdf(file_name: str, raw_text: str, pages_note: str) -> dict:
    if raw_text.startswith("__PDF_READ_ERROR__"):
        return {
            "Файл": file_name,
            "Наименование застройщика": "",
            "Общая площадь здания": "",
            "Дата ввода в эксплуатацию": "",
            "Обработанные страницы": pages_note,
            "Примечание": raw_text,
        }

    text = normalize_text(raw_text)

    if len(text) < 50:
        return {
            "Файл": file_name,
            "Наименование застройщика": "",
            "Общая площадь здания": "",
            "Дата ввода в эксплуатацию": "",
            "Обработанные страницы": pages_note,
            "Примечание": "Текст не найден. Возможно, PDF является сканом без текстового слоя или выбранные страницы не содержат текста.",
        }

    developer = extract_developer_name(text)
    area = extract_building_area(text)
    date = extract_commissioning_date(text)

    notes = []

    if not developer:
        notes.append("Не найдено наименование застройщика")
    if not area:
        notes.append("Не найдена общая площадь здания")
    if not date:
        notes.append("Не найдена дата ввода в эксплуатацию")

    return {
        "Файл": file_name,
        "Наименование застройщика": developer,
        "Общая площадь здания": area,
        "Дата ввода в эксплуатацию": date,
        "Обработанные страницы": pages_note,
        "Примечание": "; ".join(notes),
    }


# =========================
# EXCEL
# =========================

def create_excel_bytes(rows: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF данные"

    headers = [
        "№",
        "Файл",
        "Наименование застройщика",
        "Общая площадь здания",
        "Дата ввода в эксплуатацию",
        "Обработанные страницы",
        "Примечание",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(rows, start=1):
        ws.append([
            i,
            row["Файл"],
            row["Наименование застройщика"],
            row["Общая площадь здания"],
            row["Дата ввода в эксплуатацию"],
            row["Обработанные страницы"],
            row["Примечание"],
        ])

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    widths = {
        "A": 6,
        "B": 45,
        "C": 60,
        "D": 25,
        "E": 25,
        "F": 35,
        "G": 60,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


# =========================
# STREAMLIT
# =========================

st.set_page_config(
    page_title="Извлечение данных из PDF",
    layout="wide"
)

st.title("Извлечение данных из PDF в Excel")

st.markdown(
    """
    Скрипт ищет в PDF:
    
    - наименование застройщика;
    - общую площадь здания;
    - дату ввода в эксплуатацию.
    """
)

source_mode = st.radio(
    "Источник PDF",
    [
        "Папка на компьютере",
        "Загрузить PDF-файлы через интерфейс",
    ],
)

st.divider()

col1, col2, col3 = st.columns(3)

with col1:
    max_files = st.number_input(
        "Сколько файлов обработать",
        min_value=1,
        value=5,
        step=1,
    )

with col2:
    start_from_file = st.number_input(
        "С какого файла начать",
        min_value=1,
        value=1,
        step=1,
        help="1 — первый файл, 2 — второй файл и т.д.",
    )

with col3:
    page_selection = st.text_input(
        "Страницы",
        value="1-3",
        help="Примеры: 1, 2 | 1-3 | 1,5-8 | (1, 2) | (1-3) | (1,5-8). Пустое поле — все страницы.",
    )

st.caption("Примеры формата страниц: `1, 2`, `1-3`, `1,5-8`, `(1, 2)`, `(1-3)`, `(1,5-8)`.")

rows = []

if source_mode == "Папка на компьютере":
    folder_path = st.text_input(
        "Путь к папке с PDF",
        value=r"C:\Users\Andrey\Desktop\pdf",
    )

    recursive = st.checkbox(
        "Искать PDF во вложенных папках",
        value=False,
    )

    run_button = st.button("Обработать PDF из папки", type="primary")

    if run_button:
        pdf_folder = Path(folder_path)

        if not pdf_folder.exists():
            st.error("Папка не найдена. Проверь путь.")
            st.stop()

        if recursive:
            all_pdf_files = sorted(pdf_folder.rglob("*.pdf"))
        else:
            all_pdf_files = sorted(pdf_folder.glob("*.pdf"))

        if not all_pdf_files:
            st.warning("В указанной папке PDF-файлы не найдены.")
            st.stop()

        start_index = int(start_from_file) - 1
        selected_pdf_files = all_pdf_files[start_index:start_index + int(max_files)]

        st.info(f"Найдено PDF: {len(all_pdf_files)}. Будет обработано: {len(selected_pdf_files)}.")

        progress = st.progress(0)

        for i, pdf_path in enumerate(selected_pdf_files, start=1):
            with st.status(f"Обрабатываю: {pdf_path.name}", expanded=False):
                raw_text, pages_note = extract_text_from_pdf_file_path(pdf_path, page_selection)
                row = process_pdf(pdf_path.name, raw_text, pages_note)
                rows.append(row)

            progress.progress(i / len(selected_pdf_files))

elif source_mode == "Загрузить PDF-файлы через интерфейс":
    uploaded_files = st.file_uploader(
        "Загрузи PDF-файлы",
        type=["pdf"],
        accept_multiple_files=True,
    )

    run_button = st.button("Обработать загруженные PDF", type="primary")

    if run_button:
        if not uploaded_files:
            st.warning("Сначала загрузи PDF-файлы.")
            st.stop()

        selected_uploaded_files = uploaded_files[
            int(start_from_file) - 1:int(start_from_file) - 1 + int(max_files)
        ]

        st.info(f"Загружено PDF: {len(uploaded_files)}. Будет обработано: {len(selected_uploaded_files)}.")

        progress = st.progress(0)

        for i, uploaded_file in enumerate(selected_uploaded_files, start=1):
            with st.status(f"Обрабатываю: {uploaded_file.name}", expanded=False):
                raw_text, pages_note = extract_text_from_uploaded_pdf(uploaded_file, page_selection)
                row = process_pdf(uploaded_file.name, raw_text, pages_note)
                rows.append(row)

            progress.progress(i / len(selected_uploaded_files))


if rows:
    st.divider()

    st.subheader("Результат")

    st.dataframe(
        rows,
        use_container_width=True,
        hide_index=True,
    )

    excel_file = create_excel_bytes(rows)

    st.download_button(
        label="Скачать Excel",
        data=excel_file,
        file_name="pdf_result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )