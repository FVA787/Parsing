import json
import re
import time
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urljoin

import fitz  # PyMuPDF
import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# БАЗОВЫЕ НАСТРОЙКИ
# ============================================================

BASE_URL = "https://xn--80az8a.xn--d1aqf.xn--p1ai"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0 Safari/537.36"
    ),
    "Accept": "application/json,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": BASE_URL + "/",
}

# Публичные фронтовые endpoint-кандидаты.
# Если наш.дом.рф поменяет структуру, чаще всего нужно будет поправить именно этот блок.
CATALOG_ENDPOINTS = [
    f"{BASE_URL}/сервисы/api/kn/object",
    f"{BASE_URL}/сервисы/api/erz/main",
]

OBJECT_JSON_ENDPOINTS = [
    f"{BASE_URL}/сервисы/api/object/{{obj_id}}",
    f"{BASE_URL}/сервисы/api/kn/object/{{obj_id}}",
    f"{BASE_URL}/сервисы/api/erz/object/{{obj_id}}",
]

OBJECT_DOC_ENDPOINTS = [
    f"{BASE_URL}/сервисы/api/object/{{obj_id}}/documents",
    f"{BASE_URL}/сервисы/api/object/{{obj_id}}/document",
    f"{BASE_URL}/сервисы/api/object/{{obj_id}}/docs",
    f"{BASE_URL}/сервисы/api/kn/object/{{obj_id}}/documents",
    f"{BASE_URL}/сервисы/api/erz/object/{{obj_id}}/documents",
]

OBJECT_HTML_URL = f"{BASE_URL}/сервисы/каталог-новостроек/объект/{{obj_id}}"

UUID_RE = re.compile(
    r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}"
)

REGIONS = {
    "Москва": "77",
    "Московская область": "50",
    "Санкт-Петербург": "78",
    "Ленинградская область": "47",
    "Краснодарский край": "23",
    "Республика Адыгея": "01",
    "Алтайский край": "22",
    "Амурская область": "28",
    "Архангельская область": "29",
    "Воронежская область": "36",
    "Иркутская область": "38",
    "Калининградская область": "39",
    "Нижегородская область": "52",
    "Новосибирская область": "54",
    "Ростовская область": "61",
    "Самарская область": "63",
    "Свердловская область": "66",
    "Татарстан": "16",
    "Тюменская область": "72",
    "Челябинская область": "74",
    "Другой регион / вручную": "",
}


# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================

def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    return session


def sanitize_filename(value: str, max_len: int = 140) -> str:
    value = str(value or "").strip()
    value = re.sub(r'[\\/:*?"<>|]+', "_", value)
    value = re.sub(r"\s+", " ", value)
    value = value.strip(" ._")
    if not value:
        value = "file"
    return value[:max_len]


def parse_any_date(value) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    s = str(value).strip()

    if not s:
        return None

    # 2026-04-30 или 2026-04-30T10:20:30
    m = re.search(r"\d{4}-\d{1,2}-\d{1,2}", s)
    if m:
        for fmt in ("%Y-%m-%d",):
            try:
                return datetime.strptime(m.group(0), fmt).date()
            except ValueError:
                pass

    # 30.04.2026 / 30-04-2026 / 30/04/2026
    m = re.search(r"\d{1,2}[./-]\d{1,2}[./-]\d{4}", s)
    if m:
        raw = m.group(0)
        for fmt in ("%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                pass

    return None


def format_date_for_path(d: date | None) -> str:
    if not d:
        return "date_not_set"
    return d.strftime("%Y-%m-%d")


def in_date_range(value, date_from: date | None, date_to: date | None) -> bool:
    d = parse_any_date(value)

    if d is None:
        return False

    if date_from and d < date_from:
        return False

    if date_to and d > date_to:
        return False

    return True


def response_is_json(resp: requests.Response) -> bool:
    ctype = resp.headers.get("Content-Type", "").lower()
    text_start = resp.text[:50].lstrip()
    return "application/json" in ctype or text_start.startswith("{") or text_start.startswith("[")


def safe_get_json(session: requests.Session, url: str, params: dict | None = None, timeout: int = 30):
    try:
        resp = session.get(url, params=params, timeout=timeout)
        if resp.status_code != 200:
            return None, f"HTTP {resp.status_code}: {resp.url}"

        if not response_is_json(resp):
            return None, f"Ответ не JSON: {resp.url}"

        return resp.json(), ""

    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


def safe_get_text(session: requests.Session, url: str, timeout: int = 30):
    try:
        resp = session.get(url, timeout=timeout)
        if resp.status_code != 200:
            return "", f"HTTP {resp.status_code}: {resp.url}"

        return resp.text, ""

    except Exception as e:
        return "", f"{type(e).__name__}: {e}"


def find_value_by_keys(obj, keys: tuple[str, ...]):
    """
    Рекурсивно ищет первое значение по набору возможных ключей.
    """
    keys_lower = {k.lower() for k in keys}

    if isinstance(obj, dict):
        for k, v in obj.items():
            if str(k).lower() in keys_lower:
                return v

        for v in obj.values():
            found = find_value_by_keys(v, keys)
            if found is not None:
                return found

    elif isinstance(obj, list):
        for item in obj:
            found = find_value_by_keys(item, keys)
            if found is not None:
                return found

    return None


def extract_items_from_catalog_json(data):
    """
    Достает список объектов из разных возможных структур ответа:
    data.list, data.items, items, _items, results, content и т.д.
    """
    if isinstance(data, list):
        if any(isinstance(x, dict) for x in data):
            return data

    if not isinstance(data, dict):
        return []

    direct_keys = [
        "list",
        "items",
        "_items",
        "results",
        "content",
        "data",
        "objects",
    ]

    for key in direct_keys:
        value = data.get(key)

        if isinstance(value, list):
            if any(isinstance(x, dict) for x in value):
                return value

        if isinstance(value, dict):
            nested = extract_items_from_catalog_json(value)
            if nested:
                return nested

    # глубокий fallback
    for value in data.values():
        nested = extract_items_from_catalog_json(value)
        if nested:
            return nested

    return []


def get_object_id(item: dict) -> str:
    value = find_value_by_keys(
        item,
        (
            "objId",
            "objectId",
            "object_id",
            "id",
            "objectCode",
        ),
    )
    return str(value).strip() if value is not None else ""


def get_object_name(item: dict) -> str:
    value = find_value_by_keys(
        item,
        (
            "objName",
            "name",
            "complexName",
            "complexShortName",
            "residentialComplexName",
            "objAddr",
            "address",
        ),
    )
    return str(value).strip() if value is not None else ""


def get_developer_name(item: dict) -> str:
    value = find_value_by_keys(
        item,
        (
            "devShortCleanNm",
            "devShortNm",
            "developerName",
            "devName",
            "orgName",
            "shortName",
            "fullName",
        ),
    )
    return str(value).strip() if value is not None else ""


def get_publication_date(item: dict):
    return find_value_by_keys(
        item,
        (
            "objPublDt",
            "objPublishDate",
            "obj_publ_dt",
            "publishDate",
            "publicationDate",
            "pdPublDt",
            "pdPublicationDate",
            "datePubl",
            "datePublication",
        ),
    )


def object_matches_region_locally(item: dict, region_name: str, region_code: str) -> bool:
    """
    Локальный fallback-фильтр по региону, если API не отфильтровал.
    Не идеален, но помогает отсечь очевидно чужие регионы.
    """
    blob = json.dumps(item, ensure_ascii=False).lower()
    region_name_l = (region_name or "").lower().strip()
    region_code_l = str(region_code or "").lower().strip()

    if region_name_l and region_name_l != "другой регион / вручную":
        if region_name_l in blob:
            return True

    if region_code_l:
        possible_tokens = [
            f'"region": "{region_code_l}"',
            f'"region":"{region_code_l}"',
            f'"regioncode": "{region_code_l}"',
            f'"regioncode":"{region_code_l}"',
            f'"objregion": "{region_code_l}"',
            f'"objregion":"{region_code_l}"',
        ]
        if any(token in blob for token in possible_tokens):
            return True

    # Если не смогли доказать, что регион чужой, лучше не отбрасывать жестко.
    return True


def dedupe_objects(items: list[dict]) -> list[dict]:
    result = []
    seen = set()

    for item in items:
        obj_id = get_object_id(item)
        if not obj_id:
            continue

        if obj_id in seen:
            continue

        seen.add(obj_id)
        result.append(item)

    return result


# ============================================================
# ПОИСК ОБЪЕКТОВ НА НАШ.ДОМ.РФ
# ============================================================

def build_catalog_params(
    offset: int,
    limit: int,
    region_code: str,
    region_param_name: str,
    obj_status: str,
):
    params = {
        "offset": offset,
        "limit": limit,
        "sortField": "obj_publ_dt",
        "sortType": "desc",
    }

    if obj_status != "Любой":
        params["objStatus"] = {
            "Строится": "0",
            "Сдан": "1",
            "Проблемный": "2",
        }.get(obj_status, "0")

    if region_code and region_param_name and region_param_name != "не передавать":
        params[region_param_name] = region_code

    return params


def fetch_catalog_objects(
    region_name: str,
    region_code: str,
    date_from: date,
    date_to: date,
    max_objects: int,
    max_scan: int,
    obj_status: str,
    region_param_name: str,
    include_unknown_dates: bool,
    progress_callback=None,
):
    session = make_session()

    found = []
    errors = []
    limit = 100

    region_param_candidates = [
        region_param_name,
        "region",
        "regionCode",
        "objRegion",
        "objRegionCode",
        "subjectRf",
        "не передавать",
    ]

    # Убираем дубли, сохраняя порядок
    region_param_candidates = list(dict.fromkeys(region_param_candidates))

    scanned = 0

    for offset in range(0, max_scan, limit):
        page_items = []
        page_error_log = []

        for endpoint in CATALOG_ENDPOINTS:
            for reg_param in region_param_candidates:
                params = build_catalog_params(
                    offset=offset,
                    limit=limit,
                    region_code=region_code,
                    region_param_name=reg_param,
                    obj_status=obj_status,
                )

                data, err = safe_get_json(session, endpoint, params=params)

                if err:
                    page_error_log.append(err)
                    continue

                items = extract_items_from_catalog_json(data)

                if items:
                    page_items = items
                    break

            if page_items:
                break

        if not page_items:
            errors.extend(page_error_log[:5])
            break

        scanned += len(page_items)

        for item in page_items:
            if not isinstance(item, dict):
                continue

            if not object_matches_region_locally(item, region_name, region_code):
                continue

            publ_dt = get_publication_date(item)

            if publ_dt is None and not include_unknown_dates:
                continue

            if publ_dt is not None and not in_date_range(publ_dt, date_from, date_to):
                continue

            found.append(item)

            if len(found) >= max_objects:
                break

        if progress_callback:
            progress_callback(min(scanned / max_scan, 1.0), len(found), scanned)

        if len(found) >= max_objects:
            break

        time.sleep(0.25)

    return dedupe_objects(found), errors


# ============================================================
# ПОИСК PDF ПРОЕКТНЫХ ДЕКЛАРАЦИЙ
# ============================================================

def normalize_file_url(url_or_uuid: str) -> str:
    value = str(url_or_uuid or "").strip()
    value = value.replace("\\/", "/").replace("\\u002F", "/")

    if not value:
        return ""

    if UUID_RE.fullmatch(value):
        return f"{BASE_URL}/api/ext/file/{value}?inline=1"

    if value.startswith("/"):
        return urljoin(BASE_URL, value)

    if value.startswith("http://") or value.startswith("https://"):
        return value

    m = UUID_RE.search(value)
    if m:
        return f"{BASE_URL}/api/ext/file/{m.group(0)}?inline=1"

    return ""


def extract_file_urls_from_text(text: str) -> list[dict]:
    """
    Ищет URL вида /api/ext/file/<uuid>.
    Если рядом есть слова "проектная декларация" — приоритет выше.
    """
    text = text.replace("\\/", "/").replace("\\u002F", "/")

    candidates = []

    pattern = re.compile(
        r"(?:https?://[^\"'\s<>]+)?/api/ext/file/[0-9a-fA-F-]{36}(?:\?[^\"'\s<>]*)?"
    )

    for m in pattern.finditer(text):
        raw_url = m.group(0)
        start = max(0, m.start() - 700)
        end = min(len(text), m.end() + 700)
        ctx = text[start:end].lower()

        is_project_declaration = (
            "проект" in ctx and "деклара" in ctx
        )

        candidates.append(
            {
                "url": normalize_file_url(raw_url),
                "source": "html_or_json_text",
                "priority": 10 if is_project_declaration else 1,
                "context": ctx[:1000],
            }
        )

    return candidates


def iter_dicts(obj):
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from iter_dicts(v)
    elif isinstance(obj, list):
        for item in obj:
            yield from iter_dicts(item)


def extract_declaration_urls_from_json(data) -> list[dict]:
    """
    Ищет в JSON документы, похожие на проектную декларацию.
    """
    candidates = []

    # 1. Прямые URL в JSON
    raw_text = json.dumps(data, ensure_ascii=False)
    candidates.extend(extract_file_urls_from_text(raw_text))

    # 2. Словари, где текст похож на "проектная декларация",
    # а внутри есть UUID файла.
    for d in iter_dicts(data):
        d_text = json.dumps(d, ensure_ascii=False).lower()

        if not ("проект" in d_text and "деклара" in d_text):
            continue

        uuids = UUID_RE.findall(json.dumps(d, ensure_ascii=False))

        for uuid in uuids:
            candidates.append(
                {
                    "url": normalize_file_url(uuid),
                    "source": "json_uuid_near_project_declaration",
                    "priority": 20,
                    "context": d_text[:1000],
                }
            )

        for possible_key in (
            "fileId",
            "fileGuid",
            "fileUuid",
            "documentFileId",
            "documentGuid",
            "docFileId",
            "id",
            "guid",
            "uuid",
            "url",
            "downloadUrl",
            "fileUrl",
        ):
            if possible_key in d:
                url = normalize_file_url(str(d.get(possible_key)))
                if url:
                    candidates.append(
                        {
                            "url": url,
                            "source": f"json_key_{possible_key}",
                            "priority": 30,
                            "context": d_text[:1000],
                        }
                    )

    # Дедупликация с сохранением максимального приоритета
    by_url = {}

    for c in candidates:
        url = c.get("url")
        if not url:
            continue

        if url not in by_url or c.get("priority", 0) > by_url[url].get("priority", 0):
            by_url[url] = c

    return sorted(by_url.values(), key=lambda x: x.get("priority", 0), reverse=True)


def find_declaration_urls_for_object(session: requests.Session, obj_id: str, raw_item: dict | None = None):
    candidates = []

    if raw_item:
        candidates.extend(extract_declaration_urls_from_json(raw_item))

    # JSON карточки объекта
    for template in OBJECT_JSON_ENDPOINTS:
        url = template.format(obj_id=obj_id)
        data, err = safe_get_json(session, url)

        if data:
            candidates.extend(extract_declaration_urls_from_json(data))

        time.sleep(0.15)

    # JSON документов объекта
    for template in OBJECT_DOC_ENDPOINTS:
        url = template.format(obj_id=obj_id)
        data, err = safe_get_json(session, url)

        if data:
            candidates.extend(extract_declaration_urls_from_json(data))

        time.sleep(0.15)

    # HTML карточки объекта
    html_url = OBJECT_HTML_URL.format(obj_id=obj_id)
    html, err = safe_get_text(session, html_url)

    if html:
        candidates.extend(extract_file_urls_from_text(html))

    # Дедупликация
    by_url = {}
    for c in candidates:
        url = c.get("url")
        if not url:
            continue

        if url not in by_url or c.get("priority", 0) > by_url[url].get("priority", 0):
            by_url[url] = c

    return sorted(by_url.values(), key=lambda x: x.get("priority", 0), reverse=True)


def download_pdf(session: requests.Session, url: str, output_path: Path) -> tuple[bool, str]:
    try:
        resp = session.get(url, timeout=60, stream=True)

        if resp.status_code != 200:
            return False, f"HTTP {resp.status_code}"

        content = resp.content

        # Иногда content-type может быть неидеальный, поэтому проверяем первые байты.
        if not content.startswith(b"%PDF"):
            ctype = resp.headers.get("Content-Type", "")
            if "pdf" not in ctype.lower():
                return False, f"Ответ не похож на PDF. Content-Type: {ctype}"

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(content)

        return True, "ok"

    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def download_declarations(
    objects: list[dict],
    output_folder: Path,
    download_all_found: bool,
    progress_callback=None,
):
    session = make_session()
    rows = []

    for i, item in enumerate(objects, start=1):
        obj_id = get_object_id(item)
        obj_name = get_object_name(item)
        developer = get_developer_name(item)
        publ_dt_raw = get_publication_date(item)
        publ_dt = parse_any_date(publ_dt_raw)

        candidates = find_declaration_urls_for_object(session, obj_id, raw_item=item)

        if not candidates:
            rows.append(
                {
                    "obj_id": obj_id,
                    "object_name": obj_name,
                    "developer": developer,
                    "publication_date": publ_dt_raw,
                    "pdf_path": "",
                    "pdf_url": "",
                    "status": "PDF проектной декларации не найден",
                    "source": "",
                }
            )
            continue

        selected = candidates if download_all_found else candidates[:1]

        for j, cand in enumerate(selected, start=1):
            suffix = f"_{j}" if len(selected) > 1 else ""
            date_part = publ_dt.strftime("%Y-%m-%d") if publ_dt else "no_date"

            filename = sanitize_filename(
                f"{date_part}_obj_{obj_id}{suffix}_{developer}_{obj_name}.pdf",
                max_len=180,
            )

            pdf_path = output_folder / filename

            ok, status = download_pdf(session, cand["url"], pdf_path)

            rows.append(
                {
                    "obj_id": obj_id,
                    "object_name": obj_name,
                    "developer": developer,
                    "publication_date": publ_dt_raw,
                    "pdf_path": str(pdf_path) if ok else "",
                    "pdf_url": cand["url"],
                    "status": status,
                    "source": cand.get("source", ""),
                }
            )

            time.sleep(0.2)

        if progress_callback:
            progress_callback(i / len(objects), i, len(objects))

    return rows


# ============================================================
# ЧТЕНИЕ PDF И ПОИСК ОБЩЕЙ ПЛОЩАДИ
# ============================================================

def parse_page_selection(page_selection: str | None, total_pages: int) -> list[int]:
    """
    Форматы:
    ""       — все страницы
    "1"      — первая страница
    "1,2"    — страницы 1 и 2
    "(1, 2)" — страницы 1 и 2
    "1-3"    — страницы 1, 2, 3
    "1,5-8"  — страницы 1 и 5-8
    """

    if not page_selection or not str(page_selection).strip():
        return list(range(total_pages))

    text = str(page_selection).strip()
    text = text.replace("(", "").replace(")", "")
    text = text.replace(" ", "")

    if not text:
        return list(range(total_pages))

    selected_pages = set()

    for part in text.split(","):
        if not part:
            continue

        if "-" in part:
            left, right = part.split("-", 1)

            if not left.isdigit() or not right.isdigit():
                raise ValueError(f"Некорректный диапазон страниц: {part}")

            start_page = int(left)
            end_page = int(right)

            if start_page > end_page:
                start_page, end_page = end_page, start_page

            for page_number in range(start_page, end_page + 1):
                page_index = page_number - 1

                if 0 <= page_index < total_pages:
                    selected_pages.add(page_index)

        else:
            if not part.isdigit():
                raise ValueError(f"Некорректный номер страницы: {part}")

            page_number = int(part)
            page_index = page_number - 1

            if 0 <= page_index < total_pages:
                selected_pages.add(page_index)

    return sorted(selected_pages)


def extract_text_from_pdf(pdf_path: Path, page_selection: str | None = "") -> tuple[str, str]:
    text_parts = []

    with fitz.open(pdf_path) as doc:
        total_pages = len(doc)
        page_indexes = parse_page_selection(page_selection, total_pages)

        for page_index in page_indexes:
            page = doc[page_index]
            text_parts.append(page.get_text("text"))

        human_pages = ", ".join(str(i + 1) for i in page_indexes)
        pages_note = f"{human_pages} из {total_pages}"

    return "\n".join(text_parts), pages_note


def normalize_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def clean_area_value(value: str) -> str:
    value = value.replace("\n", " ")
    value = re.sub(r"\s+", " ", value)
    value = value.strip(" :-–—;,.")
    return value


AREA_PATTERNS = [
    {
        "name": "Общая площадь здания",
        "regex": r"Общая\s+площадь\s+здан(?:ия|ий)\s*[:\-–—]?\s*([\d\s.,]+)\s*(?:кв\.?\s*м|м2|м²)?",
    },
    {
        "name": "Общая площадь объекта капитального строительства",
        "regex": r"Общая\s+площадь\s+объекта\s+капитального\s+строительства\s*[:\-–—]?\s*([\d\s.,]+)\s*(?:кв\.?\s*м|м2|м²)?",
    },
    {
        "name": "Площадь объекта капитального строительства",
        "regex": r"Площадь\s+объекта\s+капитального\s+строительства\s*[:\-–—]?\s*([\d\s.,]+)\s*(?:кв\.?\s*м|м2|м²)?",
    },
    {
        "name": "Общая площадь строящегося объекта",
        "regex": r"Общая\s+площадь\s+строящ(?:егося|ихся)\s+(?:объекта|объектов)\s*[:\-–—]?\s*([\d\s.,]+)\s*(?:кв\.?\s*м|м2|м²)?",
    },
    {
        "name": "Площадь введенного объекта",
        "regex": r"(?:Площадь|Общая\s+площадь)\s+введ[её]нного\s+объекта\s*[:\-–—]?\s*([\d\s.,]+)\s*(?:кв\.?\s*м|м2|м²)?",
    },
    {
        "name": "Общая площадь с контекстом объекта",
        "regex": r"(?:объект[а-я\s]{0,80}|здани[еяй][а-я\s]{0,80})Общая\s+площадь\s*[:\-–—]?\s*([\d\s.,]+)\s*(?:кв\.?\s*м|м2|м²)?",
    },
]


def get_context(text: str, start: int, end: int, chars: int = 350) -> str:
    left = max(0, start - chars)
    right = min(len(text), end + chars)

    fragment = text[left:right]
    fragment = fragment.replace("\n", " ")
    fragment = re.sub(r"\s+", " ", fragment)

    return fragment.strip()


def extract_building_area(text: str) -> tuple[str, str, str]:
    normalized = normalize_text(text)

    for item in AREA_PATTERNS:
        regex = item["regex"]
        match = re.search(regex, normalized, flags=re.IGNORECASE | re.DOTALL)

        if match:
            value = clean_area_value(match.group(1))
            context = get_context(normalized, match.start(), match.end())
            return value, item["name"], context

    return "", "", ""


def extract_declaration_date(text: str) -> str:
    normalized = normalize_text(text)

    patterns = [
        r"Дата\s+подачи\s+декларации\s*[:\-–—]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{4})",
        r"по\s+состоянию\s+на\s*(\d{1,2}[./-]\d{1,2}[./-]\d{4})",
        r"Проектная\s+декларация.*?от\s*(\d{1,2}[./-]\d{1,2}[./-]\d{4})",
    ]

    for pattern in patterns:
        match = re.search(pattern, normalized, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(1)

    return ""


def read_existing_declarations(
    folder: Path,
    region_name: str,
    date_from: date,
    date_to: date,
    max_files: int,
    page_selection: str,
    recursive: bool,
    include_unknown_dates: bool,
    progress_callback=None,
):
    if recursive:
        pdf_files = sorted(folder.rglob("*.pdf"))
    else:
        pdf_files = sorted(folder.glob("*.pdf"))

    if max_files:
        pdf_files = pdf_files[:max_files]

    rows = []

    for i, pdf_path in enumerate(pdf_files, start=1):
        try:
            raw_text, pages_note = extract_text_from_pdf(pdf_path, page_selection)
            text = normalize_text(raw_text)

            if len(text) < 50:
                rows.append(
                    {
                        "Регион": region_name,
                        "Файл": pdf_path.name,
                        "Путь": str(pdf_path),
                        "Дата декларации": "",
                        "Общая площадь здания": "",
                        "Шаблон": "",
                        "Страницы": pages_note,
                        "Контекст": "",
                        "Примечание": "Текст почти не извлекся. Возможно, PDF является сканом.",
                    }
                )
                continue

            decl_date = extract_declaration_date(text)
            parsed_decl_date = parse_any_date(decl_date)

            if parsed_decl_date is None and not include_unknown_dates:
                continue

            if parsed_decl_date is not None:
                if date_from and parsed_decl_date < date_from:
                    continue
                if date_to and parsed_decl_date > date_to:
                    continue
            area, pattern_name, context = extract_building_area(text)   # ЗДЕСЬ ВЫЗЫВАЕТСЯ ФУНКЦИЯ ЧТЕНИЯ PDF

            note = ""
            if not area:
                note = "Общая площадь здания не найдена"

            rows.append(
                {
                    "Регион": region_name,
                    "Файл": pdf_path.name,
                    "Путь": str(pdf_path),
                    "Дата декларации": decl_date,
                    "Общая площадь здания": area,
                    "Шаблон": pattern_name,  #ЗДЕСЬ В КОЛОНКУ EXCEL ЗАПИСЫВАЕТСЯ НАЙДЕННОЕ ЗНАЧЕНИЕ ПЛОЩАДИ
                    "Страницы": pages_note,
                    "Контекст": context,
                    "Примечание": note,
                }
            )

        except Exception as e:
            rows.append(
                {
                    "Регион": region_name,
                    "Файл": pdf_path.name,
                    "Путь": str(pdf_path),
                    "Дата декларации": "",
                    "Общая площадь здания": "",
                    "Шаблон": "",
                    "Страницы": "",
                    "Контекст": "",
                    "Примечание": f"{type(e).__name__}: {e}",
                }
            )

        if progress_callback:
            progress_callback(i / len(pdf_files), i, len(pdf_files))

    return rows


# ============================================================
# EXCEL
# ============================================================

def normalize_excel_output_path(output_path: Path) -> Path:
    """
    Если пользователь указал папку, автоматически добавляем имя Excel-файла.
    Если пользователь указал путь без .xlsx, тоже добавляем .xlsx.
    """
    output_path = Path(output_path)
    # Если путь существует и это папка
    if output_path.exists() and output_path.is_dir():
        return output_path / "nash_dom_rf_area_result.xlsx"
    # Если путь заканчивается без расширения
    if output_path.suffix.lower() != ".xlsx":
        return output_path / "nash_dom_rf_area_result.xlsx"
    return output_path

def save_rows_to_excel(rows: list[dict], output_path: Path, sheet_name: str = "Результат"):
    output_path = normalize_excel_output_path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    try:
        df.to_excel(output_path, index=False, sheet_name=sheet_name)
    except PermissionError:
        raise PermissionError(
            f"Не удалось сохранить Excel-файл: {output_path}\n"
            f"Возможные причины:\n"
            f"1. Ты указал папку вместо файла .xlsx.\n"
            f"2. Excel-файл уже открыт в Excel.\n"
            f"3. Нет прав на запись в эту папку.\n\n"
            f"Попробуй закрыть Excel или указать другой путь, например:\n"
            f"{output_path.parent / 'result_new.xlsx'}"
        )
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = str(cell.value or "")
            max_len = max(max_len, min(len(value), 80))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 70))
    wb.save(output_path)
    return output_path


# ============================================================
# STREAMLIT UI
# ============================================================

st.set_page_config(
    page_title="Наш.дом.рф — проектные декларации",
    layout="wide",
)

if "mode" not in st.session_state:
    st.session_state.mode = "download"


def set_mode(mode: str):
    st.session_state.mode = mode


st.title("Сервис выгрузки и чтения проектных деклараций наш.дом.рф")

left_col, right_col = st.columns([1, 3], gap="large")

with left_col:
    st.subheader("Меню")

    st.button(
        "Выгрузить декларации",
        use_container_width=True,
        on_click=set_mode,
        args=("download",),
    )

    st.button(
        "Прочитать существующие декларации",
        use_container_width=True,
        on_click=set_mode,
        args=("read",),
    )

    st.button(
        "Записать в Базу данных",
        use_container_width=True,
        on_click=set_mode,
        args=("db",),
    )

    st.button(
        "Свободная кнопка",
        use_container_width=True,
        on_click=set_mode,
        args=("free",),
    )

    st.divider()
    st.caption("Левая четверть — навигация. Правая часть — настройки выбранного действия.")


with right_col:
    mode = st.session_state.mode

    # ========================================================
    # РЕЖИМ 1: ВЫГРУЗИТЬ ДЕКЛАРАЦИИ
    # ========================================================

    if mode == "download":
        st.subheader("Выгрузить декларации")

        with st.form("download_form"):
            col_a, col_b = st.columns(2)

            with col_a:
                region_name = st.selectbox(
                    "Регион выгрузки",
                    options=list(REGIONS.keys()),
                    index=list(REGIONS.keys()).index("Москва"),
                )

            with col_b:
                custom_region_code = st.text_input(
                    "Код региона вручную, если выбран другой регион",
                    value="",
                    help="Например: 77 — Москва, 50 — Московская область, 23 — Краснодарский край.",
                )

            region_code = REGIONS.get(region_name, "")
            if region_name == "Другой регион / вручную":
                region_code = custom_region_code.strip()

            col_d1, col_d2 = st.columns(2)

            with col_d1:
                date_from = st.date_input(
                    "Дата декларации / публикации от",
                    value=date(2026, 1, 1),
                    format="DD.MM.YYYY",
                )

            with col_d2:
                date_to = st.date_input(
                    "Дата декларации / публикации до",
                    value=date.today(),
                    format="DD.MM.YYYY",
                )

            output_base_folder = st.text_input(
                "Папка для сохранения PDF",
                value=str(Path.home() / "Desktop" / "nash_dom_rf_declarations"),
            )

            col_m1, col_m2, col_m3 = st.columns(3)

            with col_m1:
                max_objects = st.number_input(
                    "Сколько деклараций скачать максимум",
                    min_value=1,
                    max_value=5000,
                    value=20,
                    step=1,
                )

            with col_m2:
                max_scan = st.number_input(
                    "Сколько объектов просканировать максимум",
                    min_value=100,
                    max_value=20000,
                    value=1000,
                    step=100,
                    help="Если по дате мало совпадений, увеличь это значение.",
                )

            with col_m3:
                obj_status = st.selectbox(
                    "Статус объекта",
                    options=["Строится", "Сдан", "Проблемный", "Любой"],
                    index=0,
                )

            with st.expander("Расширенные настройки API"):
                region_param_name = st.selectbox(
                    "Имя параметра региона для API",
                    options=[
                        "region",
                        "regionCode",
                        "objRegion",
                        "objRegionCode",
                        "subjectRf",
                        "не передавать",
                    ],
                    index=0,
                    help="Если выгрузка не находит объекты, попробуй objRegionCode или не передавать.",
                )

                include_unknown_dates = st.checkbox(
                    "Включать объекты без найденной даты публикации",
                    value=False,
                )

                download_all_found = st.checkbox(
                    "Скачивать все найденные PDF проектных деклараций по объекту",
                    value=False,
                    help="По умолчанию скачивается первый наиболее вероятный PDF.",
                )

            submitted = st.form_submit_button(
                "Начать выгрузку деклараций",
                type="primary",
            )

        if submitted:
            if not region_code:
                st.error("Не указан код региона.")
                st.stop()

            if date_from > date_to:
                st.error("Дата 'от' не может быть позже даты 'до'.")
                st.stop()

            target_folder = (
                Path(output_base_folder)
                / sanitize_filename(f"{region_code}_{region_name}")
                / f"{format_date_for_path(date_from)}__{format_date_for_path(date_to)}"
            )

            st.info(f"Папка выгрузки: {target_folder}")

            st.write("### 1. Поиск объектов")

            progress = st.progress(0)
            status_box = st.empty()

            def search_progress(p, found_count, scanned_count):
                progress.progress(p)
                status_box.write(
                    f"Просканировано объектов: {scanned_count}. "
                    f"Подходит под фильтр: {found_count}."
                )

            objects, errors = fetch_catalog_objects(
                region_name=region_name,
                region_code=region_code,
                date_from=date_from,
                date_to=date_to,
                max_objects=int(max_objects),
                max_scan=int(max_scan),
                obj_status=obj_status,
                region_param_name=region_param_name,
                include_unknown_dates=include_unknown_dates,
                progress_callback=search_progress,
            )

            if errors:
                with st.expander("Ошибки / диагностические сообщения API"):
                    for err in errors[:50]:
                        st.write(err)

            if not objects:
                st.warning(
                    "Объекты не найдены. Попробуй увеличить 'Сколько объектов просканировать', "
                    "сменить параметр региона в расширенных настройках или поставить 'не передавать'."
                )
                st.stop()

            preview_rows = []

            for item in objects:
                preview_rows.append(
                    {
                        "ID объекта": get_object_id(item),
                        "Объект": get_object_name(item),
                        "Застройщик": get_developer_name(item),
                        "Дата публикации": get_publication_date(item),
                    }
                )

            st.success(f"Найдено объектов под выгрузку: {len(objects)}")
            st.dataframe(preview_rows, use_container_width=True, hide_index=True)

            st.write("### 2. Скачивание PDF")

            progress2 = st.progress(0)
            status_box2 = st.empty()

            def download_progress(p, current, total):
                progress2.progress(p)
                status_box2.write(f"Скачивание: {current} из {total}")

            rows = download_declarations(
                objects=objects,
                output_folder=target_folder,
                download_all_found=download_all_found,
                progress_callback=download_progress,
            )

            log_path = target_folder / "download_log.xlsx"
            save_rows_to_excel(rows, log_path, sheet_name="Выгрузка")

            st.success(f"Выгрузка завершена. Лог сохранен: {log_path}")
            st.dataframe(rows, use_container_width=True, hide_index=True)

    # ========================================================
    # РЕЖИМ 2: ПРОЧИТАТЬ СУЩЕСТВУЮЩИЕ ДЕКЛАРАЦИИ
    # ========================================================

    elif mode == "read":
        st.subheader("Прочитать существующие декларации")

        with st.form("read_form"):
            col_a, col_b = st.columns(2)

            with col_a:
                region_name_read = st.selectbox(
                    "Регион",
                    options=list(REGIONS.keys()),
                    index=list(REGIONS.keys()).index("Москва"),
                    key="read_region",
                )

            with col_b:
                folder_to_read = st.text_input(
                    "Папка с уже скачанными PDF",
                    value=str(Path.home() / "Desktop" / "nash_dom_rf_declarations"),
                )

            col_d1, col_d2 = st.columns(2)

            with col_d1:
                read_date_from = st.date_input(
                    "Дата декларации от",
                    value=date(2026, 1, 1),
                    format="DD.MM.YYYY",
                    key="read_date_from",
                )

            with col_d2:
                read_date_to = st.date_input(
                    "Дата декларации до",
                    value=date.today(),
                    format="DD.MM.YYYY",
                    key="read_date_to",
                )

            col_r1, col_r2, col_r3 = st.columns(3)

            with col_r1:
                max_files_to_read = st.number_input(
                    "Сколько PDF прочитать максимум",
                    min_value=1,
                    max_value=10000,
                    value=50,
                    step=1,
                )

            with col_r2:
                page_selection = st.text_input(
                    "Страницы для чтения",
                    value="",
                    help="Пусто — все страницы. Примеры: 1,2 | 1-3 | 1,5-8 | (1, 2).",
                )

            with col_r3:
                recursive_read = st.checkbox(
                    "Искать PDF во вложенных папках",
                    value=True,
                )

            include_unknown_decl_dates = st.checkbox(
                "Включать PDF без найденной даты декларации",
                value=True,
            )

            output_excel_path = st.text_input(
                "Куда сохранить Excel",
                value=str(Path.home() / "Desktop" / "nash_dom_rf_area_result.xlsx"),
            )

            submitted_read = st.form_submit_button(
                "Прочитать PDF и сформировать Excel",
                type="primary",
            )

        if submitted_read:
            folder = Path(folder_to_read)

            if not folder.exists():
                st.error("Папка с PDF не найдена.")
                st.stop()

            if read_date_from > read_date_to:
                st.error("Дата 'от' не может быть позже даты 'до'.")
                st.stop()

            pdf_count = len(list(folder.rglob("*.pdf"))) if recursive_read else len(list(folder.glob("*.pdf")))

            if pdf_count == 0:
                st.warning("В указанной папке PDF-файлы не найдены.")
                st.stop()

            st.info(f"Найдено PDF-файлов: {pdf_count}")

            progress = st.progress(0)
            status_box = st.empty()

            def read_progress(p, current, total):
                progress.progress(p)
                status_box.write(f"Чтение PDF: {current} из {total}")

            rows = read_existing_declarations(
                folder=folder,
                region_name=region_name_read,
                date_from=read_date_from,
                date_to=read_date_to,
                max_files=int(max_files_to_read),
                page_selection=page_selection,
                recursive=recursive_read,
                include_unknown_dates=include_unknown_decl_dates,
                progress_callback=read_progress,
            )

            if not rows:
                st.warning("После фильтрации по датам подходящих PDF не осталось.")
                st.stop()

            output_path = Path(output_excel_path)
            saved_path = save_rows_to_excel(rows, output_path, sheet_name="Площади")

            st.success(f"Excel сформирован: {saved_path}")
            st.dataframe(rows, use_container_width=True, hide_index=True)

            # Таким образом, можно указывать в интерфейсе просто папку:
            # C:\Users\..............\Desktop\Проекты\2026-05-12 Робот дом РФ - пример
            # А программа сохранит файл сюда:
            # C:\Users\fomichevv4\Desktop\Проекты\2026-05-12 Робот дом РФ\nash_dom_rf_area_result.xlsx

    # ========================================================
    # РЕЖИМ 3: БАЗА ДАННЫХ
    # ========================================================

    elif mode == "db":
        st.subheader("Записать в Базу данных")
        st.info("Эта кнопка пока зарезервирована. Позже сюда можно подключить SQLite / PostgreSQL.")

        st.code(
            """
# Будущая логика:
# 1. Выбрать Excel или папку с PDF.
# 2. Прочитать данные.
# 3. Записать строки в таблицу БД.
# 4. Сделать защиту от дублей по obj_id / pdf_url / имени файла.
            """.strip()
        )

    # ========================================================
    # РЕЖИМ 4: СВОБОДНАЯ КНОПКА
    # ========================================================

    elif mode == "free":
        st.subheader("Свободная кнопка")
        st.info("Эта кнопка пока свободна. Можно использовать под тест regex, сверку с БД или массовую переобработку PDF.")

        st.code(
            """
# Возможные будущие функции:
# - тест регулярных выражений на одном PDF;
# - просмотр сырого текста PDF;
# - повторное чтение только ошибочных файлов;
# - проверка дублей;
# - экспорт в CSV / JSON.
            """.strip()
        )